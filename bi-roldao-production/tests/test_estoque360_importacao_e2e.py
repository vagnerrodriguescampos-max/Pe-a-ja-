from __future__ import annotations

from datetime import date
from pathlib import Path
import sys

import duckdb
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from backend.estoque_contratos import TIPO_ESTOQUE, TIPO_RUPTURA
from backend.estoque_importacao import (
    _iterar_linhas_aba,
    processar_arquivo_estoque_360,
)

POS = date(2026, 8, 31)
REFERENCIA = date(2026, 9, 3)


def _criar_estoque(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Estoque"
    ws.append([
        "Loja",
        "Produto",
        "Descrição",
        "Estoque Disponível - Qtde",
        "Estoque Disponível - R$",
        "Venda Qtde - 31 DD",
        "Venda CMV",
    ])
    ws.append(["R002", 110.0, "Produto A", 5, 50, 31, 30])
    wb.save(path)


def _criar_ruptura(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "BD"
    # A base real possui linhas antes do cabeçalho; o importador não pode assumir linha 1.
    ws.append(["Relatório de Ruptura"])
    ws.append([None])
    ws.append([
        "Loja",
        "Cód Material",
        "Material",
        "Itens Ativos",
        "Itens c/ Ruptura",
        "Estoque Qtde",
        "Ped. Aberto - DL",
        "Regional",
    ])
    # Reproduz literalmente o padrão de chave encontrado na base real.
    ws.append(["R002 ROLDÃO FREGUESIA DO", 110.0, "Produto A", 1, 1, 0, 0, "INTERIOR"])
    # Reproduz ruptura que pode existir sem correspondente na base de Estoque.
    ws.append(["R007 ROLDÃO TESTE", 999.0, "Produto somente ruptura", 1, 1, 0, 0, "INTERIOR"])
    wb.save(path)


def test_iter_rows_tem_prioridade_e_nao_materializa_to_python():
    class AbaFake:
        def iter_rows(self):
            yield ["a", "b"]
            yield [1, 2]

        def to_python(self):
            raise AssertionError("to_python não deve ser chamado quando iter_rows existe")

    assert list(_iterar_linhas_aba(AbaFake())) == [["a", "b"], [1, 2]]


def test_pipeline_arquivo_calamine_staging_duckdb_e_view_unificada(tmp_path: Path):
    estoque = tmp_path / "Estoque - Venda - 31.08.xlsx"
    ruptura = tmp_path / "31.08 - Ruptura.xlsx"
    _criar_estoque(estoque)
    _criar_ruptura(ruptura)

    con = duckdb.connect(":memory:")
    try:
        r_est = processar_arquivo_estoque_360(
            con,
            caminho=estoque,
            tipo=TIPO_ESTOQUE,
            usuario="teste-e2e",
            referencia_data=REFERENCIA,
        )
        r_rup = processar_arquivo_estoque_360(
            con,
            caminho=ruptura,
            tipo=TIPO_RUPTURA,
            usuario="teste-e2e",
            referencia_data=REFERENCIA,
        )

        assert r_est["status"] == "SUCESSO"
        assert r_rup["status"] == "SUCESSO"
        assert r_est["data_posicao"] == POS.isoformat()
        assert r_rup["data_posicao"] == POS.isoformat()
        assert r_est["linhas_validas"] == 1
        assert r_rup["linhas_validas"] == 2

        # A chave extensa/numericamente diferente deve casar com a chave canônica do Estoque.
        casado = con.execute(
            """
            SELECT loja, sku, tem_estoque, tem_ruptura, item_ativo, ruptura, regional
            FROM vw_estoque_360
            WHERE loja='R002' AND sku='110'
            """
        ).fetchone()
        assert casado == ("R002", "110", True, True, True, True, "INTERIOR")

        # O FULL OUTER JOIN deve preservar ruptura real mesmo sem linha correspondente no Estoque.
        somente_ruptura = con.execute(
            """
            SELECT loja, sku, tem_estoque, tem_ruptura, item_ativo, ruptura
            FROM vw_estoque_360
            WHERE loja='R007' AND sku='999'
            """
        ).fetchone()
        assert somente_ruptura == ("R007", "999", False, True, True, True)

        assert con.execute("SELECT COUNT(*) FROM estoque_diario WHERE data_posicao=?", [POS]).fetchone()[0] == 1
        assert con.execute("SELECT COUNT(*) FROM ruptura_diaria WHERE data_posicao=?", [POS]).fetchone()[0] == 2
        assert con.execute("SELECT COUNT(*) FROM vw_estoque_360 WHERE data_posicao=?", [POS]).fetchone()[0] == 2
    finally:
        con.close()
